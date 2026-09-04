import openpyxl

wb = openpyxl.load_workbook(r'd:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx')
ws = wb['items']

lines = []
for r in range(2, min(ws.max_row+1, 200)):
    name = ws.cell(r, 7).value
    desc_vi = ws.cell(r, 8).value
    desc_cn = ws.cell(r, 3).value
    if desc_vi:
        lines.append(f'Name: {name}')
        lines.append(f'CN: {desc_cn}')
        lines.append(f'VI: {desc_vi}')
        lines.append('-'*40)

with open(r'd:\BaiTapCode\WHMX\WhmxCalc\tools\extract_desc.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

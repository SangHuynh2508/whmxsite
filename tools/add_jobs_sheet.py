import openpyxl
from pathlib import Path
import json

excel_path = Path(r'd:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx')

# Load existing workbook
wb = openpyxl.load_workbook(excel_path)

# Check if jobs sheet exists
sheet_name = 'jobs'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(title=sheet_name)
ws.append(["id", "name_cn", "name_vi"])

# Jobs data from jobs.json
jobs_data = {
    "1": {"name_cn": "宿卫", "name_vi": "Túc Vệ"},
    "2": {"name_cn": "轻锐", "name_vi": "Khinh Duệ"},
    "3": {"name_cn": "远击", "name_vi": "Viễn Kích"},
    "4": {"name_cn": "构术", "name_vi": "Cấu Thuật"},
    "5": {"name_cn": "战略", "name_vi": "Chiến Lược"}
}

for jid, data in jobs_data.items():
    ws.append([jid, data['name_cn'], data['name_vi']])

# Save
wb.save(excel_path)
print("Added jobs sheet successfully.")

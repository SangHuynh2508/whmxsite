import json
import openpyxl
from pathlib import Path

MASTER = Path(__file__).resolve().parent.parent.parent / "NeoArtifacts" / "MasterData" / "json"
EXCEL_PATH = Path(__file__).resolve().parent.parent / "localization" / "names_vi.xlsx"
PUBLIC_DIR = Path(__file__).resolve().parent.parent / "public"
OUT_FILE = PUBLIC_DIR / "data.json"

def load_json(name):
    return json.loads((MASTER / name).read_text(encoding="utf-8"))

def build():
    PUBLIC_DIR.mkdir(parents=True, exist_ok=True)
    
    print("Loading localization...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    def get_sheet(wb, possible_names, default_idx=0):
        for name in possible_names:
            if name in wb.sheetnames:
                return wb[name]
        if len(wb.worksheets) > default_idx:
            return wb.worksheets[default_idx]
        return wb.active

    char_loc = {}
    ws_char = get_sheet(wb, ["characters", "Khán Giả", "Khan Gia"], 0)
    for r in range(2, ws_char.max_row + 1):
        cid = ws_char.cell(r, 1).value
        if cid:
            char_loc[cid] = {
                "name_vi": ws_char.cell(r, 8).value or "",
                "fullname_vi": ws_char.cell(r, 9).value or "",
                "nickname_vi": ws_char.cell(r, 10).value or "",
                "tags_vi": ws_char.cell(r, 11).value or ""
            }
            
    item_loc = {}
    ws_item = get_sheet(wb, ["items", "Vật Phẩm", "Vat Pham"], 1)
    for r in range(2, ws_item.max_row + 1):
        iid = str(ws_item.cell(r, 1).value)
        if iid:
            item_loc[iid] = {
                "name_vi": ws_item.cell(r, 7).value or "",
                "category": ws_item.cell(r, 6).value or "other"
            }
            
    node_loc = {}
    ws_node = get_sheet(wb, ["talent_nodes"], 2)
    if ws_node:
        for r in range(2, ws_node.max_row + 1):
            if ws_node.max_column >= 6:
                bank_id = str(ws_node.cell(r, 2).value or "").strip()
                name_cn = str(ws_node.cell(r, 3).value or "").strip()
                name_vi = str(ws_node.cell(r, 5).value or "").strip()
                desc_vi = str(ws_node.cell(r, 6).value or "").strip()
                entry = {"name_vi": name_vi, "desc_vi": desc_vi}
                if bank_id and (name_vi or desc_vi):
                    node_loc[bank_id] = entry
                if name_cn and (name_vi or desc_vi):
                    node_loc[name_cn] = entry
            else:
                ncn = str(ws_node.cell(r, 1).value or "").strip()
                if ncn:
                    node_loc[ncn] = {
                        "desc_vi": str(ws_node.cell(r, 2).value or "").strip(),
                        "name_vi": str(ws_node.cell(r, 3).value or "").strip()
                    }

    skill_loc = {}
    ws_skill = get_sheet(wb, ["Kỹ Năng", "skills"], 3)
    if ws_skill:
        for r in range(2, ws_skill.max_row + 1):
            sk_id = str(ws_skill.cell(r, 1).value or "")
            name_vi = str(ws_skill.cell(r, 7).value or "")
            desc_vi = str(ws_skill.cell(r, 8).value or "")
            if sk_id:
                skill_loc[sk_id] = {"name_vi": name_vi, "desc_vi": desc_vi}

    profile_loc = {}
    ws_prof = get_sheet(wb, ["Hồ Sơ", "profile"], 4)
    if ws_prof:
        for r in range(2, ws_prof.max_row + 1):
            pid = str(ws_prof.cell(r, 1).value or "")
            text_vi = str(ws_prof.cell(r, 5).value or "")
            if pid:
                profile_loc[pid] = text_vi

    # EXP books that may not be in the excel but exist in game
    EXP_BOOKS = {
        "2101": {"exp": 500,  "name_vi": "Sơ Cấp Xã Hội Học"},
        "2102": {"exp": 1000, "name_vi": "Trung Cấp Xã Hội Học"},
        "2103": {"exp": 2000, "name_vi": "Cao Cấp Xã Hội Học"},
        "2104": {"exp": 4000, "name_vi": "Tiến Cấp Xã Hội Học"},
        "2105": {"exp": 8000, "name_vi": "Thâm Độ Xã Hội Học"},
    }
    # Categories for items not in excel but used in talent costs - override map
    CATEGORY_OVERRIDES = {str(i): "skill_material" for i in range(1011, 1016)}
    CATEGORY_OVERRIDES.update({str(i): "skill_material" for i in range(1001, 1006)})

    print("Processing items...")
    item_map = load_json("itemMap.json")
    items_db = {}
    for iid, data in item_map.items():
        if iid in item_loc or iid == "3" or iid in EXP_BOOKS:
            category = CATEGORY_OVERRIDES.get(iid, item_loc.get(iid, {}).get("category", "other"))
            if iid in EXP_BOOKS:
                category = "exp_book"
            items_db[iid] = {
                "id": iid,
                "name_cn": data.get("nameLanText", ""),
                "name_vi": item_loc.get(iid, {}).get("name_vi", "") or EXP_BOOKS.get(iid, {}).get("name_vi", ""),
                "category": category,
                "exp": EXP_BOOKS[iid]["exp"] if iid in EXP_BOOKS else None,
                "icon": f"assets/items/itemicon_{iid}.png"
            }

    print("Processing EXP curve...")
    level_up = load_json("characterLevelUp.json")
    exp_curve = {}
    for lv, exp in level_up.items():
        exp_curve[int(lv)] = int(exp)

    import re, unicodedata

    print("Processing skill maps & role attributes...")
    char_table_raw = load_json("characterTable.json")
    skill_map_raw = load_json("skillMap.json")
    char_skill_raw = load_json("characterSkillMap.json")
    passive_map_raw = load_json("characterPassiveSkillMap.json")
    skins_raw = load_json("characterSkins.json")
    roleattr_raw = load_json("roleattrMap.json")
    char_files_raw = load_json("characterFiles.json")
    char_file_text_raw = load_json("characterFileTextMap.json")
    relics_raw = load_json("historicalRelicsMap.json")

    DEPT_KEYWORDS = {
        "商业部": "Bộ Thương Mại",
        "资料部": "Bộ Tư Liệu",
        "技术部": "Bộ Kỹ Thuật",
        "执行部": "Bộ Hành Chính",
        "航海家": "Liên Minh Hàng Hải",
        "塞纳回廊": "Hành Lang Seine",
        "不列颠": "Học Viện Anh Quốc",
        "方塔": "Liên Minh Tháp Phương",
        "繁星花": "Hiệp Hội Hoa Phồn Tinh"
    }

    STAFF_STATUS_MAP = {
        "已登记": "Đã đăng ký",
        "待登记": "Chờ đăng ký"
    }

    STORE_STATUS_MAP = {
        "安全": "An toàn",
        "观察": "Theo dõi",
        "特勤": "Đặc cần"
    }

    def extract_char_profile(cid):
        finfo = char_files_raw.get(cid, {})
        card_intro = finfo.get("cardIntrolanText", "").strip()

        dept = ""
        for k, v in DEPT_KEYWORDS.items():
            if k in card_intro:
                dept = v
                break

        staff_raw = finfo.get("stafflanText", "").strip()
        staff_status = STAFF_STATUS_MAP.get(staff_raw, staff_raw)

        store_raw = finfo.get("storelanText", "").strip()
        entity_status = STORE_STATUS_MAP.get(store_raw, store_raw)

        record_id = finfo.get("recordID", "").strip()

        reports = []
        basic_file_ids = finfo.get("basicFileID", [])
        if isinstance(basic_file_ids, list):
            for r_id in basic_file_ids:
                r_obj = char_file_text_raw.get(r_id, {})
                title = r_obj.get("titleLanText", "").strip()
                content = r_obj.get("textLanText", "").strip()
                if title or content:
                    reports.append({
                        "id": r_id,
                        "title": title,
                        "content": content
                    })

        relic_entry = relics_raw.get(cid, {})
        relic_info = {}
        if relic_entry:
            relic_info = {
                "relic_name": relic_entry.get("relicslanText", "").strip(),
                "dynasty": relic_entry.get("dynastylanText", "").strip(),
                "museum": relic_entry.get("museumlanText", "").strip(),
                "intro": relic_entry.get("introductionlanText", "").strip()
            }

        return {
            "record_id": record_id,
            "department": dept,
            "staff_status": staff_status,
            "entity_status": entity_status,
            "eval_intro": card_intro,
            "reports": reports,
            "relic_info": relic_info
        }

    def extract_char_stats(cid):
        key = f"{cid}0"
        entry = roleattr_raw.get(key)
        if not entry or not isinstance(entry.get("Attr"), list):
            return None
        
        attrs = {}
        for item in entry["Attr"]:
            if isinstance(item, list) and len(item) > 0 and isinstance(item[0], str):
                parts = item[0].split(",")
                if len(parts) >= 3 and parts[2].isdigit():
                    attrs[parts[0]] = int(parts[2])
        
        hp_base = attrs.get("Hp", 0)
        hp_grow = attrs.get("Hp_GROW", 0)
        atk_base = attrs.get("Atk", 0)
        atk_grow = attrs.get("Atk_GROW", 0)
        pdef_base = attrs.get("PhysicDef", 0)
        pdef_grow = attrs.get("PhysicDef_GROW", 0)
        mdef_base = attrs.get("MagicDef", 0)
        mdef_grow = attrs.get("MagicDef_GROW", 0)
        
        return {
            "max_level": 120,
            "hp_base": hp_base,
            "hp_max": hp_base + hp_grow * 119,
            "atk_base": atk_base,
            "atk_max": atk_base + atk_grow * 119,
            "def_physic_base": pdef_base,
            "def_physic_max": pdef_base + pdef_grow * 119,
            "def_magic_base": mdef_base,
            "def_magic_max": mdef_base + mdef_grow * 119,
            "speed": attrs.get("Speed", 0),
            "mov": attrs.get("Mov", 0),
            "crit": attrs.get("Critical", 0),
            "crit_dmg": attrs.get("CritDmg", 150)
        }

    all_skills = {}
    for src in [char_skill_raw, passive_map_raw, skill_map_raw]:
        if isinstance(src, dict):
            for k, v in src.items():
                gid = v.get("GroupId")
                lvl = v.get("Level", v.get("level", 1))
                if gid:
                    key = (str(gid), int(lvl))
                    if key not in all_skills:
                        all_skills[key] = v
                    elif v.get("Attr") and not all_skills[key].get("Attr"):
                        all_skills[key] = v

    def parse_attr(attr_list):
        attr_dict = {}
        if not isinstance(attr_list, list): return attr_dict
        for item in attr_list:
            raw_str = ""
            if isinstance(item, list) and len(item) > 0:
                raw_str = item[0]
            elif isinstance(item, str):
                raw_str = item
            
            if "," in raw_str:
                parts = raw_str.split(",")
                if len(parts) >= 3 and parts[2] != "":
                    attr_dict[parts[0]] = parts[2]
                elif len(parts) >= 1:
                    attr_dict[parts[0]] = parts[-1]
        return attr_dict

    def resolve_desc(skill_obj):
        if not skill_obj: return ""
        raw_desc = skill_obj.get("DescriptionLanText", "")
        attr_dict = parse_attr(skill_obj.get("Attr", []))
        
        def replacer(m):
            pname = m.group(1)
            if pname in attr_dict:
                return str(attr_dict[pname])
            return m.group(0)
        
        clean_desc = re.sub(r'\[([A-Za-z0-9_]+),\d+\]?', replacer, raw_desc)
        clean_desc = re.sub(r'\{Buff_[^}]+\}', '', clean_desc)
        return clean_desc.strip()

    print("Processing talent tree...")
    talent_bank = load_json("talentBankMap.json")
    talent_map = load_json("characterTalentMap.json")
    char_talents = {}
    for tid, node in talent_map.items():
        cid = node.get("roleid")
        if not cid:
            cid = tid[:-2]
            
        if cid not in char_talents:
            char_talents[cid] = []
            
        cost = []
        if isinstance(node.get("requireCoin"), list):
            for entry in node["requireCoin"]:
                if isinstance(entry, dict) and "id" in entry:
                    cost.append({"id": str(entry["id"]), "count": entry.get("count", 0)})
        if isinstance(node.get("requireItems"), list):
            for entry in node["requireItems"]:
                if isinstance(entry, dict) and "id" in entry:
                    cost.append({"id": str(entry["id"]), "count": entry.get("count", 0)})
                    
        req_level = node.get("requireLevel", 1)
        req_talent = node.get("requireTalent", [])
        if isinstance(req_talent, str):
            req_talent = [req_talent]
            
        bank_id = node.get("talentBankId", "")
        bank_info = talent_bank.get(bank_id, {})
        icon_name = bank_info.get("icon", "")
        desc_cn = bank_info.get("DescriptionLanText", "")
        p1 = bank_info.get("talentParam1")
        p2 = bank_info.get("talentParam2")
        
        name_cn = node.get("namelanText", node.get("name", ""))
        loc_entry = node_loc.get(name_cn, {})
        name_vi = loc_entry.get("name_vi", "") if isinstance(loc_entry, dict) else loc_entry
        desc_vi = loc_entry.get("desc_vi", "") if isinstance(loc_entry, dict) else ""

        # Skill & Passive Icon Mapping
        skill_meta = None
        if isinstance(p1, str) and p1 in ["skill1", "skill2", "skill3", "skill4", "skill5", "skill6"]:
            char_raw = char_table_raw.get(cid, {})
            skill_info = char_raw.get(p1)
            if skill_info and isinstance(skill_info, list) and len(skill_info) > 0:
                gid = str(skill_info[0])
                upgraded_lvl = int(p2) if p2 else 1
                prev_lvl = max(1, upgraded_lvl - 1)
                
                s_up = all_skills.get((gid, upgraded_lvl)) or all_skills.get((gid, 1))
                s_prev = all_skills.get((gid, prev_lvl)) if upgraded_lvl > 1 else None

                if s_up:
                    s_icon_name = s_up.get("SkillIcon", "")
                    skill_meta = {
                        "skill_id": gid,
                        "skill_icon": f"assets/skills/{s_icon_name}.png" if s_icon_name else "",
                        "skill_name_cn": s_up.get("NameLanText", ""),
                        "level": upgraded_lvl,
                        "prev_level": prev_lvl if s_prev else None,
                        "desc_cn": resolve_desc(s_up),
                        "prev_desc_cn": resolve_desc(s_prev) if s_prev else ""
                    }
        
        char_talents[cid].append({
            "id": tid,
            "name_cn": name_cn,
            "name_vi": name_vi,
            "desc_vi": desc_vi,
            "desc_cn": desc_cn,
            "icon": f"assets/talents/{icon_name}.png" if icon_name else "",
            "req_level": req_level,
            "req_talent": req_talent,
            "cost": cost,
            "skill_meta": skill_meta
        })

    print("Processing rank up...")
    rank_up_raw = load_json("characterRankUpMap.json")
    rank_up_rules = {}
    for rare, levels in rank_up_raw.items():
        if not isinstance(levels, dict):
            continue
        rule_list = []
        for star in sorted([int(k) for k in levels.keys() if k.isdigit()]):
            sdata = levels[str(star)]
            mat = sdata.get("matItem", {})
            req_coin = sdata.get("coin", 0)
            rule_list.append({
                "star": star,
                "itemId": str(mat.get("id", "")),
                "count": mat.get("count", 0),
                "coin": req_coin
            })
        rank_up_rules[rare] = rule_list

    print("Processing character cards...")
    cards_dir = PUBLIC_DIR / "assets" / "cards"
    char_cards = {}
    if cards_dir.exists():
        for f in cards_dir.glob("*.png"):
            cid = f.name[:5]
            if cid not in char_cards:
                char_cards[cid] = []
            char_cards[cid].append(f.name)
        for cid in char_cards:
            char_cards[cid].sort()

    print("Processing character drawings & skins...")
    drawings_dir = PUBLIC_DIR / "assets" / "drawings"
    char_skins_processed = {}
    for cid, cskins in skins_raw.items():
        if not isinstance(cskins, list): continue
        valid_skins = []
        for sk in cskins:
            sid = sk.get("skinID", "")
            img_webp = drawings_dir / f"{sid}.webp"
            img_png = drawings_dir / f"{sid}.png"
            if sid and (img_webp.exists() or img_png.exists()):
                ext = "webp" if img_webp.exists() else "png"
                valid_skins.append({
                    "skinID": sid,
                    "name_cn": sk.get("skinNamelanText", sk.get("skinName", "")),
                    "name_vi": "Ảnh Gốc" if sk.get("bIsBaseSkin") else sk.get("skinNamelanText", "Trang Phục"),
                    "is_base": bool(sk.get("bIsBaseSkin")),
                    "description_cn": sk.get("getdescriptionLanText", ""),
                    "image": f"assets/drawings/{sid}.{ext}"
                })
        if valid_skins:
            char_skins_processed[cid] = valid_skins

    TYPE_MAP = {
        1: "Thường",
        2: "Chủ Động",
        3: "Nội Tại",
        4: "Tuyệt Kỹ",
        5: "Nội Tại",
        6: "Nội Tại",
        11: "Chiến Thuật"
    }

    print("Processing character skills...")
    char_skills = {}
    for cid, data in char_table_raw.items():
        skills_list = []
        for i in range(1, 7):
            sinfo = data.get(f"skill{i}")
            if not sinfo or not isinstance(sinfo, list) or len(sinfo) == 0:
                continue
            gid = str(sinfo[0])
            
            levels = []
            seen_levels = set()
            for src in [char_skill_raw, passive_map_raw, skill_map_raw]:
                if not isinstance(src, dict): continue
                for sk_key, sk_val in src.items():
                    if str(sk_val.get("GroupId")) == gid:
                        lvl = int(sk_val.get("Level", sk_val.get("level", 1)))
                        if lvl in seen_levels: continue
                        seen_levels.add(lvl)
                        
                        sk_id = f"{gid}_{lvl}"
                        sk_entry = skill_loc.get(sk_id, {})
                        stype = int(sk_val.get("type", 1)) if str(sk_val.get("type", "")).isdigit() else 1
                        icon_name = sk_val.get("SkillIcon", "")
                        
                        levels.append({
                            "level": lvl,
                            "name_cn": sk_val.get("NameLanText", ""),
                            "name_vi": sk_entry.get("name_vi", ""),
                            "desc_vi": sk_entry.get("desc_vi", ""),
                            "desc_cn": resolve_desc(sk_val),
                            "desc_raw": sk_val.get("DescriptionLanText", ""),
                            "icon": f"assets/skills/{icon_name}.png" if icon_name else "",
                            "type_id": stype,
                            "type": TYPE_MAP.get(stype, "Kỹ Năng"),
                            "select_range": str(sk_val.get("SelectRange", "")),
                            "effect_range": sk_val.get("EffectRange", 0),
                            "effect_range_type": str(sk_val.get("EffectRangeType", ""))
                        })
            levels.sort(key=lambda x: x["level"])
            if levels:
                skills_list.append({
                    "slot": f"skill{i}",
                    "group_id": gid,
                    "max_level": len(levels),
                    "levels": levels
                })
        char_skills[cid] = skills_list

    def slugify(text):
        if not text: return ""
        text = unicodedata.normalize("NFD", text)
        text = "".join(c for c in text if unicodedata.category(c) != "Mn")
        text = text.replace("đ", "d").replace("Đ", "d")
        text = re.sub(r'[^a-zA-Z0-9\s-]', '', text)
        text = re.sub(r'[\s-]+', '-', text).strip('-').lower()
        return text

    EXCLUDED_CHARACTER_IDS = {"W0021", "ES013"}
    chars_db = {}
    slug_counts = {}

    for cid in char_talents:
        char_talents[cid] = sorted(char_talents[cid], key=lambda x: x["id"])

    for cid, data in char_table_raw.items():
        talents = char_talents.get(cid, [])
        if cid.startswith("SCJ") or cid in EXCLUDED_CHARACTER_IDS or len(talents) == 0:
            continue

        rare = data.get("rare", 3)
        name_vi = char_loc.get(cid, {}).get("name_vi", "")
        name_cn = data.get("namelanText", data.get("name", ""))
        base_name = name_vi or name_cn or cid

        slug = slugify(base_name)
        if not slug:
            slug = cid.lower()
        if slug in slug_counts:
            slug_counts[slug] += 1
            slug = f"{slug}-{slug_counts[slug]}"
        else:
            slug_counts[slug] = 1

        chars_db[cid] = {
            "id": cid,
            "slug": slug,
            "name_cn": name_cn,
            "fullname_cn": data.get("FullnameLanText", ""),
            "name_vi": name_vi,
            "fullname_vi": char_loc.get(cid, {}).get("fullname_vi", ""),
            "nickname_vi": char_loc.get(cid, {}).get("nickname_vi", ""),
            "tags_vi": char_loc.get(cid, {}).get("tags_vi", ""),
            "rare": rare,
            "job": data.get("job", 0),
            "attacktype": data.get("attacktype", 0),
            "is_limited": bool(data.get("Linkage")),
            "icon": f"assets/avatars/{cid}.png",
            "cards": char_cards.get(cid, [f"{cid}001.png"]),
            "skins": char_skins_processed.get(cid, []),
            "talents": talents,
            "skills": char_skills.get(cid, []),
            "stats": extract_char_stats(cid),
            "profile": extract_char_profile(cid),
            "rankUpRule": rare
        }
        
    web_data = {
        "characters": chars_db,
        "items": items_db,
        "expCurve": exp_curve,
        "rankUpRules": rank_up_rules
    }
    
    OUT_FILE.write_text(json.dumps(web_data, ensure_ascii=False, separators=(',', ':')), encoding="utf-8")
    print(f"Generated {OUT_FILE} ({len(chars_db)} chars, {len(items_db)} items)")

if __name__ == "__main__":
    build()

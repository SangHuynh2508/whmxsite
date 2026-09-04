import openpyxl
from pathlib import Path

tag_dict = {
    "近战": "Cận Chiến",
    "远程": "Viễn Chiến",
    "输出": "Sát Thương",
    "承伤": "Đỡ Đòn",
    "支援": "Hỗ Trợ",
    "控制": "Khống Chế",
    "治疗": "Trị Liệu",
    "形态切换": "Chuyển Trạng Thái",
    "职业切换": "Chuyển Nghề Nghiệp",
    "技能伤害": "Sát Thương Kỹ Năng",
    "范围伤害": "Sát Thương Diện Rộng",
    "爆发": "Bạo Phát",
    "常击": "Đánh Thường",
    "免伤": "Miễn Thương",
    "自愈": "Tự Hồi Phục",
    "击退": "Đẩy Lùi",
    "召唤": "Triệu Hồi",
    "增益": "Buff",
    "嘲讽": "Khiêu Khích",
    "回击": "Phản Công",
    "突入": "Đột Kích",
    "援护": "Hộ Vệ",
    "格挡": "Đỡ Gạt",
    "构素护盾": "Lá Chắn Cấu Thuật",
    "闪避": "Né Tránh",
    "反弹伤害": "Phản Sát Thương",
    "传送": "Dịch Chuyển",
    "减益": "Debuff",
    "持续伤害": "Sát Thương Duy Trì",
    "生命吸取": "Hút Máu",
    "真实伤害": "Sát Thương Chuẩn",
    "连续行动": "Hành Động Liên Tục",
    "连击": "Liên Kích",
    "回溯": "Quay Ngược Thời Gian",
    "反击增益": "Buff Phản Công",
    "驱散": "Giải Trừ",
    "伤害穿透": "Xuyên Giáp",
    "伤害追加": "Sát Thương Kèm Theo",
    "减速": "Giảm Tốc",
    "回能": "Hồi Năng Lượng",
    "多段伤害": "Sát Thương Nhiều Đoạn",
    "射程加强": "Tăng Tầm Bắn",
    "再移动": "Di Chuyển Thêm",
    "追击": "Truy Kích",
    "集火": "Tập Trung Hỏa Lực"
}

def translate_tags():
    excel_path = Path(r"d:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["characters"]
    
    count = 0
    for r in range(2, ws.max_row + 1):
        tags_cn = ws.cell(r, 7).value
        if tags_cn:
            tags = [t.strip() for t in tags_cn.split(";") if t.strip()]
            tags_vi = []
            for t in tags:
                tags_vi.append(tag_dict.get(t, t))
            
            translated = "; ".join(tags_vi)
            ws.cell(r, 11).value = translated
            count += 1
            
    wb.save(excel_path)
    print(f"Translated tags for {count} characters.")

if __name__ == "__main__":
    translate_tags()

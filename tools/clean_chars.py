import openpyxl
from pathlib import Path

def clean_characters():
    excel_path = Path(r"d:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["characters"]
    
    rows_to_delete = []
    
    # Iterate backwards when planning to delete rows
    for r in range(ws.max_row, 1, -1):
        cid = str(ws.cell(r, 1).value or "")
        
        # Mark SCJ for deletion
        if cid.startswith("SCJ"):
            rows_to_delete.append(r)
            
        # Translate A0001
        if cid == "A0001":
            ws.cell(r, 8).value = "Nhà Sưu Tầm"
            ws.cell(r, 9).value = "Nhà Sưu Tầm"
            
    # Delete rows
    for r in rows_to_delete:
        ws.delete_rows(r)
        
    wb.save(excel_path)
    print(f"Deleted {len(rows_to_delete)} SCJ characters.")
    print("Translated A0001 as 'Nhà Sưu Tầm'.")

if __name__ == "__main__":
    clean_characters()

import openpyxl
from pathlib import Path

# Mapping exact Chinese descriptions to Vietnamese
translations = {
    "松烟、桐油烟、漆烟等制成，颜色墨黑，历史悠久，比较容易获得。": "Làm từ bồ hóng, khói dầu, bồ hóng sơn v.v., màu đen sẫm, lịch sử lâu đời, khá dễ tìm.",
    "人造聚合颜料，用途广泛，因为有沉淀现象，使用前需要摇晃。性价比较高。": "Bột màu tổng hợp, công dụng rộng rãi. Có hiện tượng lắng cặn nên cần lắc trước khi dùng. Hiệu quả sử dụng cao.",
    "从漆树割取的天然树液，髹漆时涂层很厚，所以用量很大。因需人工制取，成本较高。": "Nhựa tự nhiên lấy từ cây sơn, lớp phủ rất dày nên tốn nhiều. Phải chiết xuất thủ công nên chi phí cao.",
    "源于天然矿石，制作工艺复杂，颜色明丽。上色后能保存很久，是修整工作的不二之选，但价格昂贵。": "Nguồn gốc từ quặng tự nhiên, chế tác phức tạp. Lên màu bền lâu, là lựa chọn số một để phục chế nhưng giá đắt đỏ.",
    "最普通的刷子工具，一般由尼龙制成，比较容易获得。": "Loại cọ phổ biến nhất, thường làm từ nilon, khá dễ tìm.",
    "由柔软的羊毛制成，细密美观，性价比较高。": "Làm từ lông cừu mềm mại, mịn đẹp, hiệu quả sử dụng cao.",
    "由山兔的项背之毫精心制成，笔锋柔韧细长，成本较高。": "Làm từ lông gáy thỏ rừng, ngòi bút mềm dẻo thon dài, chi phí cao.",
    "由有经验的匠人用树棕手工编扎而成，十分好用，历史悠久，非常珍贵。": "Được thợ thủ công bện từ sợi cọ, rất dễ dùng, lịch sử lâu đời, vô cùng quý giá.",
    "最普通的粘合剂，只要有面粉和明矾就可以制作，十分容易获得。": "Chất kết dính phổ biến nhất, chỉ cần bột mì và phèn chua là làm được, rất dễ tìm.",
    "由环氧树脂和固化剂组成，粘结强度较高，是较为廉价的粘合剂。": "Gồm nhựa epoxy và chất đóng rắn, độ dính cao, là chất kết dính khá rẻ.",
    "看起来是金色的小珠子，由动物骨皮等组织提炼而成，有很长的应用历史，制作成本略高。": "Hạt nhỏ màu vàng, chiết xuất từ da xương động vật, lịch sử lâu đời, chi phí hơi cao.",
    "由海藻胶和石花菜制作而成，是近些年新研发出的粘合剂，对器物伤害很小，但研发成本很高。": "Làm từ keo rong biển và rau câu, ít gây hại cho hiện vật, nhưng chi phí nghiên cứu cao.",
    "含有铁元素，一般呈现出黄褐色、灰白色，可塑性好，比较容易获得。": "Chứa sắt, thường có màu vàng nâu hoặc trắng xám, tính dẻo tốt, khá dễ tìm.",
    "模料的一种，可以用来制造熔模，用途广泛且较为廉价。": "Một loại vật liệu dùng để đúc khuôn, ứng dụng rộng rãi và khá rẻ.",
    "可用来制作陶瓷母模及绘画基底，制作成本略高。": "Dùng làm khuôn gốc gốm sứ và lớp nền hội họa, chi phí hơi cao.",
    "可浇筑成模具，凝固后弹性较好，较为昂贵。": "Có thể đổ thành khuôn đúc, sau khi đông đặc có độ đàn hồi tốt, khá đắt đỏ.",
    "最普通的修整材料，只要有水就能使用。十分容易获得。": "Vật liệu chỉnh sửa phổ biến nhất, chỉ cần có nước là dùng được. Rất dễ tìm.",
    "亮晶晶的砂状颗粒，适合精细打磨及抛光，是应用广泛的廉价磨料。": "Hạt cát lấp lánh, thích hợp để mài dũa mịn và đánh bóng, ứng dụng rộng rãi, giá rẻ.",
    "在抛光时可以用到的高性价比磨料，但成本略高。": "Vật liệu mài hiệu quả cao dùng khi đánh bóng, chi phí hơi cao.",
    "半透明的小蜡珠，有些可爱，可以用来打磨，但价格昂贵。": "Hạt sáp nhỏ bán trong suốt, có thể dùng để đánh bóng, giá đắt đỏ.",
    "制作瓷器必须的基础材料，比较容易获得。": "Vật liệu cơ bản không thể thiếu để làm gốm sứ, khá dễ tìm.",
    "铜矿石提炼熔化后的液体，是青铜器的制作原料，用途广泛且价格低廉。": "Quặng đồng nấu chảy, nguyên liệu làm đồ đồng thanh, ứng dụng rộng rãi, giá rẻ.",
    "纹样很多，种类丰富，自古以来被广泛应用的书画装裱材料，但制作成本较高。": "Hoa văn phong phú, dùng để bồi tranh chữ từ thời xưa, chi phí cao.",
    "种类繁多的珍贵矿石，在镶嵌时会用到，但价格昂贵。": "Đủ loại khoáng thạch quý giá, dùng khi khảm nạm, giá đắt đỏ.",
    "暂未完成解析，敬请期待。": "Tài liệu đang được phân tích, vui lòng chờ đợi.",
    "用于替代初始特出器者晋升。\n特制的红色授权卡，是器者实力的证明。": "Dùng thay thế cho Khí Giả Siêu Việt khi thăng cấp.\nThẻ ủy quyền màu đỏ đặc chế, là minh chứng cho thực lực của Khí Giả.",
    "用于替代初始优异器者晋升。\n特制的金色授权卡，是器者实力的证明。": "Dùng thay thế cho Khí Giả Ưu Tú khi thăng cấp.\nThẻ ủy quyền màu vàng đặc chế, là minh chứng cho thực lực của Khí Giả."
}

def translate_missing():
    excel_path = Path(r"d:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["items"]
    
    count = 0
    for r in range(2, ws.max_row + 1):
        desc_cn = ws.cell(r, 3).value
        desc_vi = ws.cell(r, 8).value
        
        # If no translation yet but we have one in dictionary
        if not desc_vi and desc_cn in translations:
            ws.cell(r, 8).value = translations[desc_cn]
            count += 1
            
    wb.save(excel_path)
    print(f"Applied {count} new missing descriptions.")

if __name__ == "__main__":
    translate_missing()

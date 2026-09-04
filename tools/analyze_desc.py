import openpyxl
from pathlib import Path

def analyze_missing():
    excel_path = Path(r"d:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["items"]
    
    missing = {}
    for r in range(2, ws.max_row + 1):
        name_vi = ws.cell(r, 7).value
        desc_cn = ws.cell(r, 3).value
        desc_vi = ws.cell(r, 8).value
        
        if name_vi and not desc_vi and desc_cn:
            missing[desc_cn] = missing.get(desc_cn, []) + [name_vi]
            
    out_lines = []
    for desc, names in missing.items():
        out_lines.append(f"CN: {desc}")
        out_lines.append(f"Items: {', '.join(names)}")
        out_lines.append("-" * 40)
        
    with open(r"d:\BaiTapCode\WHMX\WhmxCalc\tools\missing_desc.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))

if __name__ == "__main__":
    analyze_missing()

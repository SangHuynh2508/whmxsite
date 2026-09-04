import openpyxl
from pathlib import Path

def extract_tags():
    excel_path = Path(r"d:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["characters"]
    
    tags_set = set()
    for r in range(2, ws.max_row + 1):
        tags_cn = ws.cell(r, 7).value
        if tags_cn:
            # tags might be separated by comma or something. 
            # In the game, they are often comma or pipe separated. Let's assume comma or just split by standard delimiters if needed.
            # I will just print the unique strings first to see the format.
            tags_set.add(tags_cn)
            
    out_lines = []
    for t in sorted(list(tags_set)):
        out_lines.append(t)
        
    with open(r"d:\BaiTapCode\WHMX\WhmxCalc\tools\tags_output.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))

if __name__ == "__main__":
    extract_tags()

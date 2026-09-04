import openpyxl
import re
from pathlib import Path

def simplify_desc(name_vi, desc_cn, desc_vi):
    if not desc_vi:
        return desc_vi
        
    # Simplify Thẻ Chứng Nhận
    if "Thẻ Chứng Nhận" in name_vi:
        # e.g., Thẻ Chứng Nhận Túc Vệ I
        match = re.search(r'Thẻ Chứng Nhận (.+) (I|II|III|IV|V)', name_vi)
        if match:
            job = match.group(1)
            roman = match.group(2)
            level = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5}.get(roman, 1)
            return f"Dùng để Khảo Hạch Khí Giả {job} và mở khóa kỹ năng. Độ phức tạp: {level}."
            
    # Simplify Giấy Chứng Nhận Tư Chất
    if "Giấy Chứng Nhận Tư Chất" in name_vi:
        match = re.search(r'Giấy Chứng Nhận Tư Chất (I|II|III|IV|V)', name_vi)
        if match:
            roman = match.group(1)
            level_str = {'I': 'đầu tiên', 'II': 'thứ hai', 'III': 'thứ ba', 'IV': 'thứ tư', 'V': 'thứ năm'}.get(roman, 'này')
            return f"Chứng nhận năng lực của Khí Giả, tài liệu quan trọng để đánh giá. Đây là lần {level_str} cần cung cấp."
            
    # Simplify Đồng tiền
    if name_vi == "Đông Cốc Tệ" or name_vi == "Tiền Đông Cốc":
        return "Đồng tiền do Quỹ Đông Cốc phát hành, sử dụng tại các khu thương mại trực thuộc Quỹ."
        
    return desc_vi

def fix_descriptions():
    excel_path = Path(r"d:\BaiTapCode\WHMX\WhmxCalc\localization\names_vi.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["items"]
    
    count = 0
    for r in range(2, ws.max_row + 1):
        name_vi = ws.cell(r, 7).value
        desc_cn = ws.cell(r, 3).value
        desc_vi = ws.cell(r, 8).value
        
        if name_vi and desc_vi:
            new_desc = simplify_desc(name_vi, desc_cn, desc_vi)
            if new_desc != desc_vi:
                ws.cell(r, 8).value = new_desc
                count += 1
                
    wb.save(excel_path)
    print(f"Fixed {count} descriptions.")

if __name__ == "__main__":
    fix_descriptions()

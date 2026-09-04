"""
tools/build_localization_excel.py
=================================
Trích xuất tên nhân vật + tên vật phẩm nâng cấp từ MasterData,
xuất ra Excel để Việt hóa.

Output: localization/names_vi.xlsx
  - Sheet "characters": id, name_cn, fullname_cn, rare, name_vi, fullname_vi
  - Sheet "items":      id, name_cn, description_cn, rare, type, category, name_vi

Usage:
    python tools/build_localization_excel.py
"""

import json
import sys
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

MASTER = Path(__file__).resolve().parent.parent.parent / "NeoArtifacts" / "MasterData" / "json"
OUT_DIR = Path(__file__).resolve().parent.parent / "localization"
OUT_FILE = OUT_DIR / "names_vi.xlsx"


def load_json(name: str):
    p = MASTER / name
    return json.loads(p.read_text(encoding="utf-8"))


def collect_upgrade_item_ids() -> set[str]:
    """Scan talent tree + rank up to find all item IDs used in upgrades."""
    ids = set()

    # --- characterTalentMap ---
    talent = load_json("characterTalentMap.json")
    for node in talent.values():
        if isinstance(node.get("requireCoin"), list):
            for entry in node["requireCoin"]:
                if isinstance(entry, dict) and "id" in entry:
                    ids.add(str(entry["id"]))
        if isinstance(node.get("requireItems"), list):
            for entry in node["requireItems"]:
                if isinstance(entry, dict) and "id" in entry:
                    ids.add(str(entry["id"]))

    # --- characterRankUpMap ---
    rank = load_json("characterRankUpMap.json")
    for rarity_data in rank.values():
        if not isinstance(rarity_data, dict):
            continue
        for star_data in rarity_data.values():
            if not isinstance(star_data, dict):
                continue
            mat = star_data.get("matItem")
            if isinstance(mat, dict) and mat.get("id"):
                ids.add(str(mat["id"]))

    # --- characterLevelUp uses coin (id=3), add explicitly ---
    ids.add("3")

    # --- EXP items (1011-1015 are known EXP books) ---
    for i in range(1011, 1016):
        ids.add(str(i))

    return ids


def guess_item_category(item_id: str, item: dict) -> str:
    """Heuristic category for the localization sheet."""
    iid = int(item_id) if item_id.isdigit() else 0
    name = item.get("nameLanText", "")
    desc = item.get("DescriptionLanText", "")

    if item_id == "3":
        return "currency"
    if 1011 <= iid <= 1015:
        return "exp_book"
    if 1021 <= iid <= 1025:
        return "talent_material_common"
    if 1031 <= iid <= 1035:
        return "talent_material_common"
    if 1041 <= iid <= 1045:
        return "talent_material_common"
    if 1051 <= iid <= 1055:
        return "talent_material_tier"
    if 1060 <= iid <= 1093:
        return "talent_material_specialty"
    if 2201 <= iid <= 2202:
        return "rank_up_fragment"
    if 2301 <= iid <= 2305:
        return "UNKNOWN"
    if "考核" in desc or "宿卫" in name:
        return "skill_material"
    if "徽章" in name:
        return "badge"

    return "other"


def build_characters_sheet(wb: openpyxl.Workbook):
    char_table = load_json("characterTable.json")
    ws = wb.active
    ws.title = "characters"

    headers = ["id", "name_cn", "fullname_cn", "rare", "job", "attacktype", "tags_cn",
               "name_vi", "fullname_vi", "nickname_vi", "tags_vi"]

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1a237e")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    vi_fill = PatternFill("solid", fgColor="FFF9C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Rarity fill colors
    rare_fills = {
        2: PatternFill("solid", fgColor="C8E6C9"),  # green
        3: PatternFill("solid", fgColor="BBDEFB"),  # blue
        4: PatternFill("solid", fgColor="E1BEE7"),  # purple
        5: PatternFill("solid", fgColor="FFE0B2"),  # orange
    }

    # Sort by rare desc, then id
    chars = sorted(char_table.values(), key=lambda c: (-c.get("rare", 0), c.get("id", "")))

    for ri, ch in enumerate(chars, 2):
        cid = ch.get("id", "")
        rare = ch.get("rare", 0)
        tags = ch.get("CharacterTagLanText", "")

        row_data = [
            cid,
            ch.get("namelanText", ch.get("name", "")),
            ch.get("FullnameLanText", ""),
            rare,
            ch.get("job", ""),
            ch.get("attacktype", ""),
            tags,
            "",  # name_vi
            "",  # fullname_vi
            "",  # nickname_vi
            "",  # tags_vi
        ]

        fill = rare_fills.get(rare)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            if fill and ci <= 7:
                cell.fill = fill
            if ci >= 8:  # Vietnamese columns
                cell.fill = vi_fill

    # Column widths
    widths = [10, 18, 30, 6, 6, 10, 25, 18, 30, 18, 25]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    return len(chars)


def build_items_sheet(wb: openpyxl.Workbook):
    item_map = load_json("itemMap.json")
    upgrade_ids = collect_upgrade_item_ids()

    ws = wb.create_sheet("items")

    headers = ["id", "name_cn", "description_cn", "rare", "type", "category",
               "name_vi", "description_vi"]

    hdr_fill = PatternFill("solid", fgColor="1b5e20")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    vi_fill = PatternFill("solid", fgColor="FFF9C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Category fill colors
    cat_fills = {
        "currency":                 PatternFill("solid", fgColor="FFF9C4"),
        "exp_book":                 PatternFill("solid", fgColor="DCEDC8"),
        "talent_material_common":   PatternFill("solid", fgColor="B3E5FC"),
        "talent_material_tier":     PatternFill("solid", fgColor="B2DFDB"),
        "talent_material_specialty": PatternFill("solid", fgColor="F0F4C3"),
        "rank_up_fragment":         PatternFill("solid", fgColor="E1BEE7"),
        "skill_material":           PatternFill("solid", fgColor="FFE0B2"),
    }

    # Build rows: only items used in upgrades
    items = []
    for item_id in sorted(upgrade_ids, key=lambda x: int(x) if x.isdigit() else 0):
        item = item_map.get(item_id)
        if not item:
            items.append({
                "id": item_id,
                "name_cn": f"⚠ NOT FOUND in itemMap",
                "description_cn": "",
                "rare": "",
                "type": "",
                "category": "MISSING",
            })
            continue

        cat = guess_item_category(item_id, item)
        desc = item.get("ConcisedescriptionLanText", "") or item.get("DescriptionLanText", "")
        # Truncate long descriptions
        if len(desc) > 120:
            desc = desc[:117] + "..."

        items.append({
            "id": item_id,
            "name_cn": item.get("nameLanText", ""),
            "description_cn": desc,
            "rare": item.get("rare", ""),
            "type": item.get("type", ""),
            "category": cat,
        })

    for ri, it in enumerate(items, 2):
        cat = it.get("category", "")
        fill = cat_fills.get(cat)

        row_data = [
            it["id"], it["name_cn"], it["description_cn"],
            it["rare"], it["type"], it["category"],
            "",  # name_vi
            "",  # description_vi
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            if fill and ci <= 6:
                cell.fill = fill
            if ci >= 7:
                cell.fill = vi_fill

    # Column widths
    widths = [10, 25, 60, 6, 6, 26, 25, 60]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    return len(items)


def build_talent_names_sheet(wb: openpyxl.Workbook):
    """Sheet for talent node name patterns (many are shared across characters)."""
    talent = load_json("characterTalentMap.json")

    # Collect unique talent name patterns
    name_set: OrderedDict[str, str] = OrderedDict()
    for node in talent.values():
        name = node.get("namelanText", node.get("name", ""))
        if name and name not in name_set:
            desc = node.get("des", "")
            name_set[name] = desc

    ws = wb.create_sheet("talent_nodes")
    headers = ["name_cn", "example_description", "name_vi"]

    hdr_fill = PatternFill("solid", fgColor="4a148c")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    vi_fill = PatternFill("solid", fgColor="FFF9C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, (name, desc) in enumerate(name_set.items(), 2):
        for ci, val in enumerate([name, desc, ""], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            if ci == 3:
                cell.fill = vi_fill

    widths = [30, 60, 30]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    return len(name_set)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading MasterData...")
    wb = openpyxl.Workbook()

    print("Building characters sheet...")
    n_chars = build_characters_sheet(wb)
    print(f"  -> {n_chars} characters")

    print("Building items sheet...")
    n_items = build_items_sheet(wb)
    print(f"  -> {n_items} upgrade items")

    print("Building talent node names sheet...")
    n_talents = build_talent_names_sheet(wb)
    print(f"  -> {n_talents} unique talent names")

    wb.save(OUT_FILE)
    print(f"\n[OK] Output: {OUT_FILE}")
    print(f"   Sheets: characters ({n_chars}), items ({n_items}), talent_nodes ({n_talents})")
    print("   -> Fill name_vi / fullname_vi / description_vi columns and save.")


if __name__ == "__main__":
    main()
